import os
import shutil
from datetime import datetime, timedelta
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.admin.models import LogEntry, ADDITION, CHANGE, DELETION
from django.contrib.contenttypes.models import ContentType
from django.db.models import Count, Q, F, Subquery, OuterRef, DateTimeField, Case, When
from django.contrib.auth.models import User
from django.contrib import messages
from django.http import JsonResponse, HttpResponseForbidden, HttpResponse
from django.conf import settings
from django.utils import timezone
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from ..models import Thesis, Category, Submission, DownloadLog, RejectedThesis, Course, Department
from authapp.models import Profile

def log_admin_action(user, obj, action_flag, message):
    """Helper to log custom admin actions."""
    from django.contrib.admin.models import LogEntry
    from django.contrib.contenttypes.models import ContentType

    LogEntry.objects.log_action(
        user_id=user.id,
        content_type_id=ContentType.objects.get_for_model(obj).pk,
        object_id=obj.pk,
        object_repr=str(obj),
        action_flag=action_flag,
        change_message=message,
    )

@login_required
@user_passes_test(lambda u: hasattr(u, 'profile') and u.profile.role in [Profile.ADMIN, Profile.LIBRARIAN])
def admin_dashboard(request):
    # Stat cards
    total_theses = Thesis.objects.count()  # all theses, including archived (archiving doesn't delete them)
    total_users = User.objects.count()
    pending_submissions = Submission.objects.filter(status='pending').count()  # live queue
    approved_theses = Submission.objects.filter(status='approved').count()     # ever approved
    download_logs = DownloadLog.objects.count()

    # Get all submissions and archives from the last 365 days in single queries
    one_year_ago = timezone.now() - timedelta(days=365)
    submissions = list(Submission.objects.filter(created_at__gte=one_year_ago).values('id', 'created_at', 'status'))
    archive_entries = list(LogEntry.objects.filter(
        action_time__gte=one_year_ago,
        change_message__icontains="Archived thesis"
    ).values('id', 'action_time'))
    today = timezone.now().date()

    # 1. 1 Month (Daily) - last 30 days
    one_month_labels = []
    one_month_approved = []
    one_month_pending = []
    one_month_rejected = []
    one_month_entered = []
    one_month_archived = []
    for i in range(29, -1, -1):
        day = today - timedelta(days=i)
        one_month_labels.append(day.strftime('%b %d'))
        day_subs = [s for s in submissions if s['created_at'].date() == day]
        one_month_approved.append(sum(1 for s in day_subs if s['status'] == 'approved'))
        one_month_pending.append(sum(1 for s in day_subs if s['status'] == 'pending'))
        one_month_rejected.append(sum(1 for s in day_subs if s['status'] == 'rejected'))
        one_month_entered.append(len(day_subs))
        one_month_archived.append(sum(1 for e in archive_entries if e['action_time'].date() == day))

    # 2. 3 Months (Weekly) - last 12 weeks
    three_month_labels = []
    three_month_approved = []
    three_month_pending = []
    three_month_rejected = []
    three_month_entered = []
    three_month_archived = []
    for i in range(11, -1, -1):
        start_date = today - timedelta(weeks=i+1)
        end_date = today - timedelta(weeks=i)
        three_month_labels.append(f"Wk {12-i}")
        week_subs = [s for s in submissions if start_date < s['created_at'].date() <= end_date]
        three_month_approved.append(sum(1 for s in week_subs if s['status'] == 'approved'))
        three_month_pending.append(sum(1 for s in week_subs if s['status'] == 'pending'))
        three_month_rejected.append(sum(1 for s in week_subs if s['status'] == 'rejected'))
        three_month_entered.append(len(week_subs))
        three_month_archived.append(sum(1 for e in archive_entries if start_date < e['action_time'].date() <= end_date))

    # 3. 6 Months (Monthly) - last 6 calendar months
    six_month_labels = []
    six_month_approved = []
    six_month_pending = []
    six_month_rejected = []
    six_month_entered = []
    six_month_archived = []
    for i in range(5, -1, -1):
        year = today.year
        month = today.month - i
        if month <= 0:
            month += 12
            year -= 1
        month_date = timezone.datetime(year, month, 1).date()
        six_month_labels.append(month_date.strftime('%b'))
        month_subs = [s for s in submissions if s['created_at'].year == year and s['created_at'].month == month]
        six_month_approved.append(sum(1 for s in month_subs if s['status'] == 'approved'))
        six_month_pending.append(sum(1 for s in month_subs if s['status'] == 'pending'))
        six_month_rejected.append(sum(1 for s in month_subs if s['status'] == 'rejected'))
        six_month_entered.append(len(month_subs))
        six_month_archived.append(sum(1 for e in archive_entries if e['action_time'].year == year and e['action_time'].month == month))

    # 4. 9 Months (Monthly) - last 9 calendar months
    nine_month_labels = []
    nine_month_approved = []
    nine_month_pending = []
    nine_month_rejected = []
    nine_month_entered = []
    nine_month_archived = []
    for i in range(8, -1, -1):
        year = today.year
        month = today.month - i
        if month <= 0:
            month += 12
            year -= 1
        month_date = timezone.datetime(year, month, 1).date()
        nine_month_labels.append(month_date.strftime('%b'))
        month_subs = [s for s in submissions if s['created_at'].year == year and s['created_at'].month == month]
        nine_month_approved.append(sum(1 for s in month_subs if s['status'] == 'approved'))
        nine_month_pending.append(sum(1 for s in month_subs if s['status'] == 'pending'))
        nine_month_rejected.append(sum(1 for s in month_subs if s['status'] == 'rejected'))
        nine_month_entered.append(len(month_subs))
        nine_month_archived.append(sum(1 for e in archive_entries if e['action_time'].year == year and e['action_time'].month == month))

    # 5. 12 Months (Monthly/Annual) - last 12 calendar months
    twelve_month_labels = []
    twelve_month_approved = []
    twelve_month_pending = []
    twelve_month_rejected = []
    twelve_month_entered = []
    twelve_month_archived = []
    for i in range(11, -1, -1):
        year = today.year
        month = today.month - i
        if month <= 0:
            month += 12
            year -= 1
        month_date = timezone.datetime(year, month, 1).date()
        twelve_month_labels.append(month_date.strftime('%b %Y') if i == 11 or month_date.month == 1 else month_date.strftime('%b'))
        month_subs = [s for s in submissions if s['created_at'].year == year and s['created_at'].month == month]
        twelve_month_approved.append(sum(1 for s in month_subs if s['status'] == 'approved'))
        twelve_month_pending.append(sum(1 for s in month_subs if s['status'] == 'pending'))
        twelve_month_rejected.append(sum(1 for s in month_subs if s['status'] == 'rejected'))
        twelve_month_entered.append(len(month_subs))
        twelve_month_archived.append(sum(1 for e in archive_entries if e['action_time'].year == year and e['action_time'].month == month))

    trend_data = {
        '1m': {'labels': one_month_labels, 'approved': one_month_approved, 'pending': one_month_pending, 'rejected': one_month_rejected, 'entered': one_month_entered, 'archived': one_month_archived},
        '3m': {'labels': three_month_labels, 'approved': three_month_approved, 'pending': three_month_pending, 'rejected': three_month_rejected, 'entered': three_month_entered, 'archived': three_month_archived},
        '6m': {'labels': six_month_labels, 'approved': six_month_approved, 'pending': six_month_pending, 'rejected': six_month_rejected, 'entered': six_month_entered, 'archived': six_month_archived},
        '9m': {'labels': nine_month_labels, 'approved': nine_month_approved, 'pending': nine_month_pending, 'rejected': nine_month_rejected, 'entered': nine_month_entered, 'archived': nine_month_archived},
        '12m': {'labels': twelve_month_labels, 'approved': twelve_month_approved, 'pending': twelve_month_pending, 'rejected': twelve_month_rejected, 'entered': twelve_month_entered, 'archived': twelve_month_archived},
    }

    theses_by_course = list(
        Thesis.objects
        .filter(course__isnull=False)          # include archived — distribution should never shrink
        .values(course_name=F('course__name'))
        .annotate(count=Count('id'))
        .order_by('-count')[:16]
    )
    for item in theses_by_course:
        name = item['course_name']
        item['course_name'] = " ".join(name.split()[4:]) if len(name.split()) > 4 else name

    theses_by_department = list(
        Thesis.objects
        .filter(department__isnull=False)      # include archived — distribution should never shrink
        .values('department__name')
        .annotate(count=Count('id'))
        .order_by('-count')[:16]
    )
    for item in theses_by_department:
        name = item['department__name']
        if " - " in name:
            item['department__name'] = name.split(" - ")[0]

    recent_theses = (
        Submission.objects
        .annotate(
            rejected_date=Subquery(
                RejectedThesis.objects.filter(original_submission_id=OuterRef('id')).values('rejected_at')[:1]
            ),
            activity_date=Case(
                When(status='approved', then='approved_at'),
                When(status='rejected', then='rejected_date'),
                default='created_at',
                output_field=DateTimeField(),
            )
        )
        .select_related('submitter', 'department', 'category', 'course')
        .order_by('-activity_date')[:16]
    )

    archived_theses_count = Thesis.objects.filter(is_archived=True).count()

    context = {
        'total_theses': total_theses,
        'total_users': total_users,
        'pending_submissions': pending_submissions,
        'approved_theses': approved_theses,
        'download_logs': download_logs,
        'theses_by_department': theses_by_department,
        'recent_theses': recent_theses,
        'theses_by_course': theses_by_course,
        'trend_data': trend_data,
        'archived_theses_count': archived_theses_count,
    }
    return render(request, 'main/admin_dashboard.html', context)

def _refine_log_data(log):
    """Helper to translate technical log entries into human-readable text."""
    action_map = {1: "Created", 2: "Updated", 3: "Removed"}
    action = action_map.get(log.action_flag, "System Task")
    details = log.change_message
    
    msg = log.change_message
    model_name = log.content_type.model.lower()
    
    if not msg:
        if log.action_flag == 1:
            action = f"New {model_name.capitalize()}"
        elif log.action_flag == 2:
            action = f"Edit {model_name.capitalize()}"
        elif log.action_flag == 3:
            action = f"Delete {model_name.capitalize()}"
        details = f"Administrative action on {model_name} record."
    elif "[APPROVED]" in msg:
        action = "Thesis Approved"
        details = msg.replace("[APPROVED]", "").strip()
    elif "[REJECTED]" in msg:
        action = "Thesis Rejected"
        details = msg.replace("[REJECTED]", "").strip()
    elif "Archived thesis" in msg:
        action = "System Archival"
        details = "Thesis record archived due to age (10+ years)."
    elif "Permanently deleted user" in msg:
        action = "Account Deletion"
    elif "Updated user role" in msg:
        action = "Role Assignment"
    elif "BULK IMPORT" in msg:
        action = "Data Import"
        details = msg.replace("BULK IMPORT:", "").strip()
    elif "Login:" in msg:
        action = "User Login"
        details = msg.replace("Login:", "").strip()
    elif "Login Attempt:" in msg:
        action = "Login Failed"
        details = msg.replace("Login Attempt:", "").strip()
    elif "Security:" in msg:
        action = "Security Event"
        details = msg.replace("Security:", "").strip()
    elif "Account Creation:" in msg:
        action = "Account Created"
        details = msg.replace("Account Creation:", "").strip()
    elif "Account Edit:" in msg:
        action = "Account Updated"
        details = msg.replace("Account Edit:", "").strip()
    elif "Account Deletion:" in msg:
        action = "Account Deleted"
        details = msg.replace("Account Deletion:", "").strip()
    elif "Login" in msg or "password" in msg:
        action = "Security Alert"

    return action, details

@login_required
@user_passes_test(lambda u: hasattr(u, 'profile') and u.profile.role in [Profile.ADMIN, Profile.LIBRARIAN])
def admin_log_entries(request):
    all_logs_qs = LogEntry.objects.all().select_related('user', 'content_type').order_by('-action_time')
    
    # Process logs for display
    def process_logs(qs):
        processed = []
        for log in qs:
            action, details = _refine_log_data(log)
            log.refined_action = action
            log.refined_details = details
            processed.append(log)
        return processed

    security_logs = process_logs(all_logs_qs.filter(
        Q(change_message__icontains='Login') | 
        Q(change_message__icontains='password') | 
        Q(change_message__icontains='reset') |
        Q(change_message__icontains='Security')
    ))
    user_logs     = process_logs(all_logs_qs.filter(
        Q(change_message__icontains='role') | 
        Q(change_message__icontains='deleted user') |
        Q(change_message__icontains='Account') |
        Q(change_message__icontains='Student') |
        Q(change_message__icontains='Professor') |
        Q(change_message__icontains='Librarian') |
        Q(change_message__icontains='Admin Staff') |
        Q(content_type__model__in=['student', 'professor', 'librarian', 'adminstaff', 'profile'])
    ))
    thesis_logs   = process_logs(all_logs_qs.filter(
        Q(change_message__icontains='APPROVED') | 
        Q(change_message__icontains='Rejected') | 
        Q(change_message__icontains='Submission') | 
        Q(change_message__icontains='Thesis') |
        Q(content_type__model__in=['thesis', 'submission', 'rejectedthesis'])
    ))
    import_logs   = process_logs(all_logs_qs.filter(change_message__icontains='BULK IMPORT'))
    all_logs      = process_logs(all_logs_qs)

    active_tab = request.GET.get('tab', 'all')
    return render(request, 'main/admin_log_entries.html', {
        'all_logs': all_logs, 'security_logs': security_logs, 'user_logs': user_logs,
        'thesis_logs': thesis_logs, 'import_logs': import_logs, 'active_tab': active_tab
    })

@login_required
@user_passes_test(lambda u: u.is_staff)
def password_reset_requests(request):
    """View active password reset codes and activation codes for offline assistance."""
    from ..models import PasswordResetCode, VerificationCode
    active_resets = PasswordResetCode.objects.filter(is_used=False, expires_at__gt=timezone.now()).select_related('user').order_by('-created_at')
    active_activations = VerificationCode.objects.filter(is_verified=False, expires_at__gt=timezone.now()).select_related('user').order_by('-created_at')
    
    return render(request, 'main/admin_reset_codes.html', {
        'active_codes': active_resets,
        'active_activations': active_activations
    })

@login_required
@user_passes_test(lambda u: hasattr(u, 'profile') and u.profile.role in [Profile.ADMIN, Profile.LIBRARIAN])
def user_list(request):
    all_users = User.objects.select_related('profile').order_by('-date_joined')
    admins     = all_users.filter(profile__role=Profile.ADMIN)
    librarians = all_users.filter(profile__role=Profile.LIBRARIAN)
    professors = all_users.filter(profile__role=Profile.PROFESSOR)
    students   = all_users.filter(profile__role=Profile.STUDENT)
    active_tab = request.GET.get('tab', 'admin')
    return render(request, 'main/user_list.html', {
        'admins': admins, 'librarians': librarians, 'professors': professors,
        'students': students, 'active_tab': active_tab, 'total_count': all_users.count(),
    })

@login_required
@user_passes_test(lambda u: hasattr(u, 'profile') and u.profile.role == Profile.ADMIN)
def delete_user(request, user_id):
    try:
        user = User.objects.get(id=user_id)
        username = user.username
        user.delete()
        log_admin_action(request.user, request.user, DELETION, f"Permanently deleted user account: {username}")
        messages.success(request, f"User account '{username}' permanently deleted.")
    except User.DoesNotExist:
        messages.warning(request, "This user account has already been deleted or does not exist.")
    return redirect('user_list')

@login_required
@user_passes_test(lambda u: hasattr(u, 'profile') and u.profile.role == Profile.ADMIN)
def edit_user(request, user_id):
    target_user = get_object_or_404(User, id=user_id)
    if request.method == 'POST':
        role = request.POST.get('role')
        profile, _ = Profile.objects.get_or_create(user=target_user)
        if role:
            profile.role = role
            profile.save()
            target_user.is_staff = (role in [Profile.ADMIN, Profile.LIBRARIAN])
            target_user.save()
            log_admin_action(request.user, target_user, CHANGE, f"Updated user role to: {role}")
        messages.success(request, "User role updated successfully.")
        return redirect('user_list')
    return render(request, 'main/edit_user.html', {'edit_user': target_user})

@login_required
@user_passes_test(lambda u: u.is_staff)
def pending_submissions(request):
    theses = Submission.objects.filter(status='pending').order_by('-created_at')
    return render(request, 'main/pending_submissions.html', {'theses': theses})

@login_required
@user_passes_test(lambda u: u.is_staff)
def approve_thesis(request, thesis_id):
    submission = get_object_or_404(Submission, id=thesis_id)
    if submission.status == Submission.STATUS_APPROVED:
        messages.info(request, f"Submission '{submission.title}' is already approved.")
        return redirect('pending_submissions')
    if request.method == 'POST':
        lc_classification = request.POST.get('lc_classification', '').strip()
        try:
            thesis = submission.approve(approved_by=request.user, lc_classification=lc_classification)
            # Cleanup auto logs
            LogEntry.objects.filter(user=request.user, content_type=ContentType.objects.get_for_model(thesis), object_id=thesis.id, action_flag=ADDITION).delete()
            LogEntry.objects.filter(user=request.user, content_type=ContentType.objects.get_for_model(submission), object_id=submission.id, action_flag__in=[CHANGE, DELETION]).delete()
            log_admin_action(request.user, thesis, ADDITION, f"[APPROVED] Submission '{submission.title}' and moved to Thesis table")
            messages.success(request, f"Submission '{submission.title}' approved successfully.")
        except ValueError as e:
            messages.error(request, f"Could not approve '{submission.title}': {str(e)}")
    return redirect('pending_submissions')

@login_required
@user_passes_test(lambda u: u.is_staff)
def reject_thesis(request, thesis_id):
    submission = get_object_or_404(Submission, id=thesis_id)
    if request.method == 'POST':
        if submission.status == Submission.STATUS_REJECTED:
            return redirect('pending_submissions')
        rejection_reason = request.POST.get('rejection_reason', '').strip()
        if not rejection_reason:
            messages.error(request, "Rejection reason is required.")
            return redirect('pending_submissions')
        try:
            rejected_thesis = submission.reject(rejection_reason=rejection_reason, rejected_by=request.user)
            # Cleanup
            LogEntry.objects.filter(user=request.user, content_type=ContentType.objects.get_for_model(rejected_thesis), object_id=rejected_thesis.id, action_flag=ADDITION).delete()
            LogEntry.objects.filter(user=request.user, content_type=ContentType.objects.get_for_model(submission), object_id=submission.id, action_flag__in=[CHANGE, DELETION]).delete()
            log_admin_action(request.user, rejected_thesis, ADDITION, f"[REJECTED] Submission '{submission.title}' - Reason: {rejection_reason[:100]}")
            messages.success(request, f"Successfully rejected '{submission.title}'.")
        except ValueError as e:
            messages.error(request, f"Could not reject: {str(e)}")
        return redirect('pending_submissions')
    return render(request, 'main/reject_thesis.html', {'submission': submission})

@login_required
@user_passes_test(lambda u: u.is_staff)
def theses_list(request):
    theses = Thesis.objects.filter(is_archived=False)
    for thesis in theses:
        thesis.course_display = str(thesis.course).split('-')[0].strip() if thesis.course else "N/A"
    return render(request, 'main/theses.html', {'theses': theses})

@login_required
@user_passes_test(lambda u: u.is_staff)
def rejected_thesis_list(request):
    rejected_theses = RejectedThesis.objects.all()
    for thesis in rejected_theses:
        thesis.course_display = str(thesis.course).split('-')[0].strip() if thesis.course else "N/A"
    return render(request, 'main/rejected_thesis.html', {'rejected_theses': rejected_theses})

@login_required
@user_passes_test(lambda u: u.is_staff)
def departments_list(request):
    if request.method == 'POST':
        # Enforce Admin Role
        if not hasattr(request.user, 'profile') or request.user.profile.role != Profile.ADMIN:
            return HttpResponseForbidden("Only administrators can add or delete colleges/departments.")
            
        admin_name = request.POST.get('admin_name', '').strip()
        action_reason = request.POST.get('action_reason', '').strip()
        action_date = request.POST.get('action_date')
        password = request.POST.get('password')

        if not password or not request.user.check_password(password):
            messages.error(request, "Authentication failed. Invalid password.")
            return redirect('departments_list')
            
        if not admin_name or not action_reason or not action_date:
            messages.error(request, "All confirmation fields (Name, Reason, Date, Password) are required.")
            return redirect('departments_list')

        name = request.POST.get('name', '').strip()
        category_id = request.POST.get('category_id')
        if name and category_id:
            category = get_object_or_404(Category, id=category_id)
            if Department.objects.filter(name=name, category=category).exists():
                messages.error(request, f"Department '{name}' already exists under '{category.name}'.")
            else:
                Department.objects.create(name=name, category=category)
                messages.success(request, f"Department '{name}' successfully created by administrator {admin_name} on {action_date}. Reason: {action_reason}")
            return redirect('departments_list')
            
    departments = Department.objects.annotate(courses_count=Count('courses'))
    categories = Category.objects.all()
    return render(request, 'main/departments.html', {
        'departments': departments,
        'categories': categories
    })

@login_required
@user_passes_test(lambda u: u.is_staff)
def courses_list(request):
    if request.method == 'POST':
        # Enforce Admin Role
        if not hasattr(request.user, 'profile') or request.user.profile.role != Profile.ADMIN:
            return HttpResponseForbidden("Only administrators can add or delete degree programs.")
            
        admin_name = request.POST.get('admin_name', '').strip()
        action_reason = request.POST.get('action_reason', '').strip()
        action_date = request.POST.get('action_date')
        password = request.POST.get('password')

        if not password or not request.user.check_password(password):
            messages.error(request, "Authentication failed. Invalid password.")
            return redirect('courses_list')
            
        if not admin_name or not action_reason or not action_date:
            messages.error(request, "All confirmation fields (Name, Reason, Date, Password) are required.")
            return redirect('courses_list')

        name = request.POST.get('name', '').strip()
        department_id = request.POST.get('department_id')
        if name and department_id:
            department = get_object_or_404(Department, id=department_id)
            if Course.objects.filter(name=name, department=department).exists():
                messages.error(request, f"Degree Program '{name}' already exists under '{department.name}'.")
            else:
                Course.objects.create(name=name, department=department)
                messages.success(request, f"Degree Program '{name}' successfully created by administrator {admin_name} on {action_date}. Reason: {action_reason}")
            return redirect('courses_list')
            
    courses = Course.objects.select_related('department').all()
    departments = Department.objects.select_related('category').all()
    for course in courses:
        course.course_display = str(course.name).split('-')[0].strip()
    return render(request, 'main/Courses.html', {
        'courses': courses,
        'departments': departments
    })

@login_required
@user_passes_test(lambda u: u.is_staff)
def admin_categories(request):
    if request.method == 'POST':
        # Enforce Admin Role
        if not hasattr(request.user, 'profile') or request.user.profile.role != Profile.ADMIN:
            return HttpResponseForbidden("Only administrators can add or delete graduation levels.")
            
        admin_name = request.POST.get('admin_name', '').strip()
        action_reason = request.POST.get('action_reason', '').strip()
        action_date = request.POST.get('action_date')
        password = request.POST.get('password')

        if not password or not request.user.check_password(password):
            messages.error(request, "Authentication failed. Invalid password.")
            return redirect('admin_categories')
            
        if not admin_name or not action_reason or not action_date:
            messages.error(request, "All confirmation fields (Name, Reason, Date, Password) are required.")
            return redirect('admin_categories')

        name = request.POST.get('name', '').strip()
        if name:
            if Category.objects.filter(name=name).exists():
                messages.error(request, f"Graduation Level '{name}' already exists.")
            else:
                Category.objects.create(name=name)
                messages.success(request, f"Graduation Level '{name}' successfully created by administrator {admin_name} on {action_date}. Reason: {action_reason}")
            return redirect('admin_categories')
            
    categories = Category.objects.all()
    category_data = []
    for category in categories:
        category_data.append({
            'id': category.id,
            'name': category.name,
            'department_count': category.departments.count(),
            'course_count': sum(dept.courses.count() for dept in category.departments.all()),
        })
    return render(request, 'main/admin_categories.html', {'categories': category_data})

@login_required
@user_passes_test(lambda u: hasattr(u, 'profile') and u.profile.role == Profile.ADMIN)
def delete_category(request, category_id):
    if request.method != 'POST':
        return HttpResponseForbidden("Invalid request method.")
        
    category = get_object_or_404(Category, id=category_id)
    name = category.name
    
    admin_name = request.POST.get('admin_name', '').strip()
    action_reason = request.POST.get('action_reason', '').strip()
    action_date = request.POST.get('action_date')
    password = request.POST.get('password')

    if not password or not request.user.check_password(password):
        messages.error(request, "Authentication failed. Invalid password.")
        return redirect('admin_categories')
        
    if not admin_name or not action_reason or not action_date:
        messages.error(request, "All confirmation fields are required.")
        return redirect('admin_categories')

    category.delete()
    messages.success(request, f"Graduation Level '{name}' has been successfully deleted by administrator {admin_name} on {action_date}. Reason: {action_reason}")
    return redirect('admin_categories')

@login_required
@user_passes_test(lambda u: hasattr(u, 'profile') and u.profile.role == Profile.ADMIN)
def delete_department(request, department_id):
    if request.method != 'POST':
        return HttpResponseForbidden("Invalid request method.")
        
    department = get_object_or_404(Department, id=department_id)
    name = department.name
    
    admin_name = request.POST.get('admin_name', '').strip()
    action_reason = request.POST.get('action_reason', '').strip()
    action_date = request.POST.get('action_date')
    password = request.POST.get('password')

    if not password or not request.user.check_password(password):
        messages.error(request, "Authentication failed. Invalid password.")
        return redirect('departments_list')
        
    if not admin_name or not action_reason or not action_date:
        messages.error(request, "All confirmation fields are required.")
        return redirect('departments_list')

    department.delete()
    messages.success(request, f"College/Department '{name}' has been successfully deleted by administrator {admin_name} on {action_date}. Reason: {action_reason}")
    return redirect('departments_list')

@login_required
@user_passes_test(lambda u: hasattr(u, 'profile') and u.profile.role == Profile.ADMIN)
def delete_course(request, course_id):
    if request.method != 'POST':
        return HttpResponseForbidden("Invalid request method.")
        
    course = get_object_or_404(Course, id=course_id)
    name = course.name
    
    admin_name = request.POST.get('admin_name', '').strip()
    action_reason = request.POST.get('action_reason', '').strip()
    action_date = request.POST.get('action_date')
    password = request.POST.get('password')

    if not password or not request.user.check_password(password):
        messages.error(request, "Authentication failed. Invalid password.")
        return redirect('courses_list')
        
    if not admin_name or not action_reason or not action_date:
        messages.error(request, "All confirmation fields are required.")
        return redirect('courses_list')

    course.delete()
    messages.success(request, f"Degree Program '{name}' has been successfully deleted by administrator {admin_name} on {action_date}. Reason: {action_reason}")
    return redirect('courses_list')

@login_required
@user_passes_test(lambda u: hasattr(u, 'profile') and u.profile.role == Profile.ADMIN)
def archive_old_theses(request):
    if request.method != "POST":
        return JsonResponse({"error": "Invalid"}, status=400)
    system_user = User.objects.filter(is_superuser=True).first()
    cutoff_year = datetime.now().year - 10
    old_theses = Thesis.objects.filter(year__lte=cutoff_year, is_archived=False)
    archived_count = 0
    errors = []

    # Permanent archive destination outside the webroot (for manual backup / cold storage)
    PERMANENT_ARCHIVE_DIR = os.path.join("C:\\", "ThesisLibrary_Permanent_Archive", "thesis_files")

    for thesis in old_theses:
        if thesis.file and thesis.file.name:
            old_path = thesis.file.path

            # 1. Move file into media/thesis_files/Archived/ (keeps it accessible in the web app)
            archive_dir = os.path.join(settings.MEDIA_ROOT, "thesis_files", "Archived")
            os.makedirs(archive_dir, exist_ok=True)
            filename = os.path.basename(old_path)
            new_path = os.path.join(archive_dir, filename)

            if os.path.exists(old_path):
                shutil.move(old_path, new_path)
                thesis.file.name = f"thesis_files/Archived/{filename}"

                # 2. Copy to C:\ThesisLibrary_Permanent_Archive\thesis_files\ for permanent off-app backup
                try:
                    os.makedirs(PERMANENT_ARCHIVE_DIR, exist_ok=True)
                    permanent_copy_path = os.path.join(PERMANENT_ARCHIVE_DIR, filename)
                    if not os.path.exists(permanent_copy_path):
                        shutil.copy2(new_path, permanent_copy_path)
                except Exception as copy_err:
                    errors.append(f"Could not copy '{filename}' to permanent archive: {str(copy_err)}")

        thesis.is_archived = True
        thesis.save()
        archived_count += 1
        log_admin_action(system_user, thesis, CHANGE, "Archived thesis (Metadata preserved)")

    response_data = {"archived": archived_count}
    if errors:
        response_data["warnings"] = errors
    return JsonResponse(response_data)
@login_required
@user_passes_test(lambda u: hasattr(u, 'profile') and u.profile.role == Profile.ADMIN)
def export_system_logs(request):
    """
    Professional Excel export for System Audit Logs.
    Supports period filtering (day, week, month, year, all) 
    and tab filtering (security, user, thesis, import, all).
    """
    period = request.GET.get('period', 'all')
    tab = request.GET.get('tab', 'all')
    
    all_logs = LogEntry.objects.select_related('user', 'content_type', 'user__profile').order_by('-action_time')
    
    # 1. Period Filtering
    now = timezone.now()
    if period == 'day':
        all_logs = all_logs.filter(action_time__gte=now - timezone.timedelta(days=1))
    elif period == 'week':
        all_logs = all_logs.filter(action_time__gte=now - timezone.timedelta(weeks=1))
    elif period == 'month':
        all_logs = all_logs.filter(action_time__gte=now - timezone.timedelta(days=30))
    elif period == 'year':
        all_logs = all_logs.filter(action_time__gte=now - timezone.timedelta(days=365))

    # 2. Tab Filtering
    if tab == 'security':
        all_logs = all_logs.filter(Q(change_message__icontains='Login') | Q(change_message__icontains='password') | Q(change_message__icontains='reset'))
    elif tab == 'user':
        all_logs = all_logs.filter(Q(change_message__icontains='role') | Q(change_message__icontains='deleted user'))
    elif tab == 'thesis':
        all_logs = all_logs.filter(Q(change_message__icontains='APPROVED') | Q(change_message__icontains='Rejected') | Q(change_message__icontains='Submission') | Q(change_message__icontains='Thesis'))
    elif tab == 'import':
        all_logs = all_logs.filter(change_message__icontains='BULK IMPORT')

    # Create Workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "System Audit Logs"

    # Define Styles
    header_fill = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin', color="DDDDDD"),
        right=Side(style='thin', color="DDDDDD"),
        top=Side(style='thin', color="DDDDDD"),
        bottom=Side(style='thin', color="DDDDDD")
    )
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Title & Metadata
    ws.merge_cells('A1:G1')
    ws['A1'] = "THESIS LIBRARY SYSTEM - AUDIT LOG REPORT"
    ws['A1'].font = Font(bold=True, size=16, color="1B5E20")
    ws['A1'].alignment = center_align
    
    ws.merge_cells('A2:G2')
    ws['A2'] = f"Report Period: {period.upper()} | Generated on: {now.strftime('%B %d, %Y %I:%M %p')}"
    ws['A2'].font = Font(italic=True, size=10)
    ws['A2'].alignment = center_align
    
    ws.append([]) # Empty row

    # Table Headers
    headers = ["Timestamp", "User", "Role", "System Module", "Specific Item", "Action Taken", "Details"]
    ws.append(headers)
    
    header_row = 4
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border

    for idx, log in enumerate(all_logs, 1):
        row_num = header_row + idx
        role = log.user.profile.get_role_display() if hasattr(log.user, 'profile') else "N/A"
        
        # Use refinement helper
        action, details = _refine_log_data(log)
        
        row_data = [
            log.action_time.strftime("%Y-%m-%d %I:%M %p"),
            log.user.username,
            role,
            log.content_type.model.capitalize(),
            log.object_repr,
            action.upper(),
            details
        ]
        
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border
            cell.alignment = left_align if col_num in [5, 7] else center_align
            
            # Alternate row colors
            if idx % 2 == 0:
                cell.fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")

    # Set column widths
    dims = {
        'A': 22, # Timestamp
        'B': 15, # User
        'C': 12, # Role
        'D': 15, # Module
        'E': 30, # Item
        'F': 15, # Action
        'G': 50  # Details
    }
    for col, value in dims.items():
        ws.column_dimensions[col].width = value

    # Freeze top rows
    ws.freeze_panes = "A5"

    # Prepare Response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"System_Log_{tab.upper()}_{period.upper()}_{now.strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    
    return response
